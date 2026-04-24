"""Функции для генерации и форматирования Excel-отчётов."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


def format_worksheet(worksheet) -> None:
    """
    Применяет елиный стиль к листу Excel:
        жирный заголовок с синим фоном, автофильтр, ширина столбцов по контенту.
    Args:
        worksheet: объект листа openpyxl (Worksheet).
    """

    # Стиль заголовка
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="305496", end_color="305496", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Применяем стиль к первой строке (заголовку)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Автофильтр на всю таблицу
    worksheet.auto_filter.ref = worksheet.dimensions

    # Ширина столбцов - по максимальной длине содержимого
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        # Прибавляем 2 для визуального запаса, ограничиваем максимум 50
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    # Закрепить первую строку (шапка всегда видна при прокрутке)
    worksheet.freeze_panes = "A2"


def build_excel_report(
    df_all: pd.DataFrame,
    df_unpaid: pd.DataFrame,
    df_vendors: pd.DataFrame,
    output_path: str,
) -> None:
    """
    Создаёт Excel-файл с тремя форматированными листами.

    Args:
        df_all: все счета.
        df_unpaid: только неоплаченные.
        df_vendors: сводка по поставщикам.
        output_path: путь к итоговому xlsx-файлу.
    """

    # Шаг 1: pandas записывает данные
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="All Invoices", index=False)
        df_unpaid.to_excel(writer, sheet_name="Unpaid", index=False)
        df_vendors.to_excel(writer, sheet_name="Vendor", index=False)

    # Шаг 2: openpyxl открывает файл и форматирует
    workbook = load_workbook(output_path)
    for sheet_name in workbook.sheetnames:
        format_worksheet(workbook[sheet_name])
    workbook.save(output_path)


if __name__ == "main":
    print("Модуль reports.py импортируется, а не запускается напрямую.")
